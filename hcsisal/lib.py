import openpyxl as xl
import os
import numpy as np

from time import strftime

from arrangement.lib import check_file_headers
from arrangement.consts import ST_ERR, ST_FIN
from hcsisal.consts import UPLOAD_DIR


def csi_recalculate(recalculate):
    """
    """
    # Проверить что шапка загруженного файла совпадает с шаблоном
    if not check_file_headers(recalculate.original_file.path, r'./hcsisal/docs/rebus_template.xlsx'):
        recalculate.processing_comments = 'Стобцы в загруженном файле не совпадают с шаблоном'
        recalculate.processing_result = ST_ERR
        return

    wb = xl.load_workbook(recalculate.original_file.path)
    ws = wb.active
    general_data = {}
    # Весы ответов
    five_scale = {'5': 100, '4': 75, '3': 50, '2': 25, '1': 0}
    RUS14_v = {'3': 0, '2': 100, '1': 100}
    RUS36_v = {'2': 100, '1': 0}
    RUS3944_v = {'2': 0, '1': 100}
    RUS41_v = {'0': 0, '1': 0, '2': 25, '3': 25, '4': 50, '5': 50, '6': 75, '7': 75, '8': 75, '9': 100, '10': 100}

    # Вопросы участвующие в рассчёте индекса c номерами колонок в файле данных и весами (кроме 5-ти бальной шкалы)
    rel_q = {
        'RUS5': {'col': 40}, 'RUS7': {'col': 42}, 'RUS8': {'col': 43}, 'RUS10': {'col': 45}, 'RUS12': {'col': 47},
        'RUS15': {'col': 50}, 'RUS16': {'col': 51}, 'RUS17': {'col': 53}, 'RUS18': {'col': 54}, 'RUS19': {'col': 55},
        'RUS20': {'col': 56}, 'RUS21': {'col': 57}, 'RUS24': {'col': 60}, 'RUS26': {'col': 62}, 'RUS28': {'col': 64},
        'RUS30': {'col': 66}, 'RUS32': {'col': 68}, 'RUS14': {'col': 49, 'veight': RUS14_v},
        'RUS36': {'col': 72, 'veight': RUS36_v}, 'RUS39': {'col': 75, 'veight': RUS3944_v},
        'RUS44': {'col': 80, 'veight': RUS3944_v}, 'RUS41': {'col': 77, 'veight': RUS41_v}
    }
    # Назначим веса для пятибальных вопросов
    for quest in rel_q.items():
        if 'veight' not in quest[1]:
            rel_q[quest[0]]['veight'] = five_scale
    # openpyxl.ws.max_row даёт неккоректный результат, т.к. считает когда-либо заполненные строки, получим руками по
    # первому столбцу
    max_row = 1
    while ws.cell(column=1, row=max_row).value:
        max_row += 1

    for row in range(2, max_row):
        dealer_code = ws.cell(column=16, row=row).value
        recode_values = []
        for question in rel_q.items():
            val = ws.cell(row=row, column=question[1]['col']).value
            recode_values.append(question[1]['veight'].get(str(val)))
        # Надо убрать None элементы
        recode_values = [value for value in recode_values if value is not None]
        if dealer_code not in general_data:
            general_data[dealer_code] = {'marks': []}
        general_data[dealer_code]['marks'].extend(recode_values)

    for dealer in general_data:
        general_data[dealer]['Index_m'] = np.average(general_data[dealer]['marks'])

    all_marks = []
    for dealer in general_data:
        all_marks.extend(general_data[dealer]['marks'])

    Index_qmean = np.average(all_marks) * 10
    for row in range(2, max_row):
        dealer_code = ws.cell(column=16, row=row).value
        Index_m = str(general_data[dealer_code]['Index_m'] * 10).replace('.', ',')
        ws.cell(column=81, row=row).value = Index_m
        ws.cell(column=82, row=row).value = Index_m
        ws.cell(column=83, row=row).value = str(Index_qmean).replace('.', ',')
    proc_file_name = os.path.join(
        UPLOAD_DIR,
        os.path.basename(
            recalculate.original_file.path).replace('.xlsx', '_CSI recalculate_%s.xlsx' % strftime("%d.%m.%Y"))
    )
    wb.save(proc_file_name)
    recalculate.csi_index = Index_qmean
    recalculate.processed_file = proc_file_name
    recalculate.processing_result = ST_FIN
