import numpy as np
import openpyxl as xl
import os
import pandas as pd

from mercedes_cli.consts import CAWI_P2_REBUS, CATI_P2_REBUS, HQ_TEMPL_P1, HQ_TEMPL_P2, REBUS_TO_HQ_P1, \
    REBUS_TO_HQ_P2, MULTIPLE_QUESTIONS, UPLOAD_MBR_DIR


def generate_gssn_reports(quarter_data):
    """
    Сгенерировать репорты для штаб-квартиры мереседес
    :param quarter_data: набор исходных файлов в виде экзмеляра модели
    :type quarter_data: mercedes_cli.models.CLIReport
    :return: Excel-файлы отчётов по первой и второй-третьей фазе
    """
    rebus = pd.read_excel(quarter_data.rebus_data.path)
    vars_for_del = ('HIDDEN_CRM_DEALER', 'CS3', 'Q_SF01', 'v321_SF01', 'v322_RM2', 'Q_RM2')
    for var in vars_for_del:
        del rebus[var]

    # Вытащить из сырых данных кави интервью по SB1=2
    cawi_p2 = pd.read_excel(quarter_data.cawi_data_p2.path)

    # Убрать тестовые и пилотные интервью
    cawi_p2 = cawi_p2[cawi_p2['Respondent.ID'].str.contains('pil') == False]
    cawi_p2_SB1_2 = cawi_p2[(cawi_p2.SB1 == 2)]

    # Добавим в данные ребуса данные кави-п2 по SB1=2
    rebus = add_data_to_rebus(rebus, cawi_p2_SB1_2, CAWI_P2_REBUS)

    # Вытащить из сырых данных кати интервью по SB1=2
    cati_p2 = pd.read_excel(quarter_data.cati_data_p2.path)
    cati_p2_SB1_2 = cati_p2[(cati_p2.SB1 == 2)]
    rebus = add_data_to_rebus(rebus, cati_p2_SB1_2, CATI_P2_REBUS)

    # Отфильтровать автомобили S-class. Добавим коды моделей
    rebus['model_code'] = np.NaN
    cawi_p1 = pd.read_excel(quarter_data.cawi_data_p1.path)
    cawi_p1 = cawi_p1[cawi_p1['Respondent.ID'].str.contains('pil') == False]
    cati_p1 = pd.read_excel(quarter_data.cati_data_p1.path)

    model_codes = get_model_codes(cati_p1, cawi_p1, cati_p2, cawi_p2)
    for row in rebus.iterrows():
        survey = int(row[1].HIDDEN_Survey)
        phase = int(row[1].HIDDEN_phase)
        interview_number = encode_rebus_id(row[1].Respondent_Serial, survey, phase)
        if phase == 3: phase = 2
        rebus.loc[row[0], 'model_code'] = model_codes[survey][phase][interview_number]
    rebus = rebus[~rebus.model_code.isin((1, 3, 49, 51, 54))]

    # Удалим интервью не-владельцев
    rebus = rebus[~rebus.VA1.isin((2, ))]

    for row in rebus.iterrows():
        # VIN для обезличенных интервью, пропуски в CIPOS, CINEG и CLI04a указываем как 9999
        for quest in ('HIDDEN_ VIN', 'CLI04a', 'CIPOS', 'CINEG'):
            if row[1][quest] is np.NaN or row[1][quest] == 999:
                rebus.loc[row[0], quest] = 9999
        # В глобальной анкете нет варианта 4 для SB1, меняем на 1
        if row[1]['SB1'] == 4.0:
            rebus.loc[row[0], 'SB1'] = 1.0
        # Перекодировать старый код Авроры-Авто
        if row[1]['Hidden_final_dealer'] == 999:
            rebus.loc[row[0], 'Hidden_final_dealer'] = 999
        for quest in MULTIPLE_QUESTIONS:
            if row[1][quest] > 0:
                rebus.loc[row[0], quest] = 1
            else:
                rebus.loc[row[0], quest] = 0

    gssn = pd.read_excel(quarter_data.gssn_file.path)
    gssn_codes = {}
    for row in gssn.iterrows():
        gssn_codes[row[1][0]] = {'code': row[1][2], 'label': row[1][3]}
    rebus['gssn_code'] = np.NaN
    rebus['gssn_label'] = np.NaN
    rebus['CountryName'] = 'anonymous'
    rebus['CountryCode'] = 999
    rebus['Quarter'] = 999
    for row in rebus.iterrows():
        rebus.loc[row[0], 'gssn_code'] = gssn_codes[row[1]['Hidden_final_dealer']]['code']
        rebus.loc[row[0], 'gssn_label'] = gssn_codes[row[1]['Hidden_final_dealer']]['label']

    r_file_p1 = xl.load_workbook(filename=HQ_TEMPL_P1, keep_vba=True)
    r_sheet_p1 = r_file_p1['CSI Sales - DATA']

    r_file_p2 = xl.load_workbook(filename=HQ_TEMPL_P2, keep_vba=True)
    r_sheet_p2 = r_file_p2['CLI - DATA']

    row_number_p1 = 7
    row_number_p2 = 7
    for row in rebus.iterrows():
        if row[1].HIDDEN_phase == 1:
            column_number = 1
            while r_sheet_p1.cell(row=6, column=column_number).value:
                head_label = r_sheet_p1.cell(row=6, column=column_number).value
                if head_label in rebus:
                    r_sheet_p1.cell(row=row_number_p1, column=column_number).value = row[1][head_label]
                else:
                    r_sheet_p1.cell(row=row_number_p1, column=column_number).value = row[1][REBUS_TO_HQ_P1[head_label]]
                column_number += 1
            row_number_p1 += 1
        else:
            column_number = 1
            while r_sheet_p2.cell(row=6, column=column_number).value:
                head_label = r_sheet_p2.cell(row=6, column=column_number).value
                if head_label in rebus:
                    r_sheet_p2.cell(row=row_number_p2, column=column_number).value = row[1][head_label]
                else:
                    r_sheet_p2.cell(row=row_number_p2, column=column_number).value = row[1][REBUS_TO_HQ_P2[head_label]]
                column_number += 1
            row_number_p2 += 1

    p1_file_name = os.path.join(UPLOAD_MBR_DIR, os.path.basename(HQ_TEMPL_P1.replace('_orig.xlsm', '.xlsm')))
    p2_file_name = os.path.join(UPLOAD_MBR_DIR, os.path.basename(HQ_TEMPL_P2.replace('_orig.xlsm', '.xlsm')))
    r_file_p1.save(p1_file_name)
    r_file_p2.save(p2_file_name)
    quarter_data.report_p1 = p1_file_name
    quarter_data.report_p2 = p2_file_name


def add_data_to_rebus(rebus, additional_source, data_map):
    """
    Добавить в данные ребуса интервью из других частей исследования
    :param rebus: Данные с портала ребус
    :type rebus: pandas.DataFrame
    :param additional_source: Кака-либо часть исследования
    :type additional_source: pandas.DataFrame
    :param data_map: Соответствие названия переменнонй источника данных к ребусу
    :return: Дололненный данные ребуса
    :return type: pandas.DataFrame
    """
    row_number = len(rebus)
    source_heads = [head for head in additional_source]
    rebus_heads = [head for head in rebus]
    for row in additional_source.iterrows():
        for var, val in zip(source_heads, row[1]):
            if var == 'INTNR':
                rebus.loc[row_number, 'Respondent_Serial'] = decode_rebus_id(
                    val, 1, row[1].HIDDEN_PHASE, row[1].HIDDEN_WAVE
                )
            elif var == 'Respondent.Serial':
                rebus.loc[row_number, 'Respondent_Serial'] = decode_rebus_id(
                    val, 2, row[1].HIDDEN_phase, row[1].HIDDEN_wave
                )
            elif var in rebus_heads:
                rebus.loc[row_number, var] = val
            elif var in data_map:
                rebus.loc[row_number, data_map[var]] = val
        row_number += 1
    return rebus


def encode_rebus_id(rebus_id, survey, phase):
    """
    Возвращает оригинальную id источника данных
    :param rebus_id: id портала ребус
    :param survey: тип проекта (кати/кави)
    :param phase: фаза исследования
    :return: original_id
    """
    int_part = int(str(int(rebus_id))[-8:])
    if survey == 1:
        if phase == 1:
            original_id = int_part
        else:
            original_id = int_part - 1000000
    else:
        if phase == 1:
            original_id = int_part - 100000
        else:
            original_id = int_part - 1100000
    return original_id


def decode_rebus_id(original_id, survey, phase, wave):
    """
    Возвращает id источника данных в формате ребуса
    :param rebus_id: id портала ребус
    :param survey: тип проекта (кати/кави)
    :param phase: фаза исследования
    :param wave: волна исследования
    :return: original_id
    """
    wave = ('0' + str(wave))[-6:]
    if survey == 1:
        if phase == 1:
            rebus_id = original_id
        else:
            rebus_id = original_id + 1000000
    else:
        if phase == 1:
            rebus_id = original_id + 100000
        else:
            rebus_id = original_id + 1100000

    rebus_id = '{year}{month}{id}'.format(year=wave[-4:], month=wave[:2], id=(('00000000' + str(rebus_id))[-8:]))
    return rebus_id


def get_model_codes(cati_p1, cawi_p1, cati_p2, cawi_p2):
    """
    Вернуть словарь содержащий коды моделей покрошенные по фазам и методдам сбора данных
    Аргументы - pandas.DataFrame
    :return: dict {survey: {phase1 : {id: ModelCode}, phase2: {id: ModelCode}}}
    """
    models = {}
    models[1] = {}
    models[1][1] = {}
    for row in cati_p1.iterrows():
        models[1][1][row[1].INTNR] = row[1].HIDDEN_MODELCODE
    models[1][2] = {}
    for row in cati_p2.iterrows():
        models[1][2][row[1].INTNR] = row[1].HIDDEN_MODELCODE

    models[2] = {}
    models[2][1] = {}
    for row in cawi_p1.iterrows():
        models[2][1][row[1][0]] = row[1].HIDDEN_Model_Code
    models[2][2] = {}
    for row in cawi_p2.iterrows():
        models[2][2][row[1][0]] = row[1].HIDDEN_Model_Code

    return models
