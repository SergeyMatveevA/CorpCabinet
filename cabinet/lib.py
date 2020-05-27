import openpyxl as xl
import pandas as pd
import re

from collections import OrderedDict
from datetime import datetime as dt, timedelta

from django.http import HttpResponse
from django.shortcuts import render
from pytils.translit import translify

from arrangement.models import Survey as Internal_Survey
from arrangement.lib import auto_fit_excel_columns, nipo_connect
from cabinet.consts import FAKE_SURVEYS, DAY_TOTAL, SURVEY_TOTAL, RESUME_STAT, CONTACT_LOG_TIME_COLUMNS, \
    FW_MONITOR_HEADERS
from cabinet.models import ResponseCode
from nipo_db.models import Survey


def get_surveys(quota=None):
    """
    Получить список активных проектов за последние 30 дней
    :return:
    """
    start_date = dt.now().date() - timedelta(days=30)
    if quota is not None:
        survey_names = [
            s.__str__()
            for s in Internal_Survey.objects.filter(last_sample_activity__gte=start_date)
            if s.cati_programm_name not in FAKE_SURVEYS
        ]
    else:
        survey_names = [
            s.__str__()
            for s in Internal_Survey.objects.filter(last_sample_activity__gte=start_date)
            if s.cati_programm_name not in FAKE_SURVEYS and s.stratification is not None
        ]
    projects = [(enum+1, project) for enum, project in enumerate(sorted(survey_names))]
    projects.insert(0, (0, 'Выберите проект'))
    return projects


def cabinet_empty_render(request, form, page_name, render_type, template, date_start=dt.now(), date_end=dt.now()):
    """
    Рендер пустой страницы для личного кабинета
    """
    resume_stat = RESUME_STAT.get(render_type, '')
    render_content = {
        'form': form, 'resume_stat': resume_stat, 'date_start': dt.strftime(date_start, '%Y-%m-%d'),
        'date_end': dt.strftime(date_end, '%Y-%m-%d'), 'page_name': page_name
    }
    return render(request, template, render_content)


def get_sample_name(selected_survey):
    """
    Получить имя таблицы в нипе
    :param selected_survey: Выбранный на форме личного кабинета проект
    """
    if Survey.objects.filter(excelsheet=selected_survey).exists():
        return 'Sample{survey}'.format(survey=Survey.objects.get(excelsheet=selected_survey).surveyname)
    else:
        return 'Sample{survey}'.format(survey=selected_survey)


def get_sample_data(survey, type):
    """
    Получить агрегированную информацию из сэмпла
    """
    sample_data = pd.read_sql(get_dialing_query(survey, type), nipo_connect())
    sample_data['date'] = sample_data[sample_data.columns[0]].dt.date
    del sample_data[sample_data.columns[0]]
    sample_data['date'].fillna('Не использовалось', inplace=True)
    return (
        sample_data,
        pd.crosstab(sample_data[sample_data.columns[0]], sample_data['date'], margins=True, margins_name='Total')
    )


def get_dialing_query(survey, type):
    """
    @param survey: Имя проекта
    @param type: Тип запроса - КонтактЛог или Сэмпл
    """
    if type == 1:
        return "select contacttime, responsecode " \
               "from {sname} order by contacttime".format(sname=get_sample_name(survey))
    elif type == 2:
        return "select contactdatetime, responsestatus from contactlog " \
               "where surveyname = '{}' order by contactdatetime".format(get_sample_name(survey).replace('Sample', ''))


def get_dates(date_start, date_end, sample_data):
    """
    Получить диапазоны дат содержащихся в сэмпле проекта и запрошенных ползователем
    @return:
    """
    check_date_range = [
        date.date() for date in pd.date_range(date_start.date(), date_end.date()).to_pydatetime().tolist()
    ]
    target_dates = [date for date in sample_data['date'].unique() if date in check_date_range]
    target_dates.insert(0, SURVEY_TOTAL)
    return target_dates


def get_excel_rows(format_data, target_dates):
    """
    Сгенерировать строки статистики для excel-файла выгрузки
    """
    excel_rows = []
    for response, date_counts in format_data[format_data.columns].transpose().to_dict().items():
        label = get_response_label(response)
        response_counts = []
        for date in target_dates:
            if type(date) == str:
                response_counts.append(date_counts.get('Total', 0))
            else:
                response_counts.append(date_counts.get(date, 0))
        excel_rows.append({label: response_counts})
    excel_rows = sorted(excel_rows, key=lambda x: list(x.keys())[0])
    return excel_rows


def get_screen_rows(format_data, target_dates):
    """
    Сгенерировать строки статистики для фронта
    """
    screen_rows = []
    for response, date_counts in format_data[format_data.columns].transpose().to_dict().items():
        label = get_response_label(response)
        response_counts = []
        for date in target_dates:
            if type(date) == str:
                response_counts.append(date_counts.get('Total', 0))
            else:
                response_counts.append(date_counts.get(date, 0))
        if len(response_counts) > 10:
            screen_rows.append({label: [response_counts[0], *response_counts[-9:]]})
        else:
            screen_rows.append({label: response_counts})
    screen_rows = sorted(screen_rows, key=lambda x: list(x.keys())[0])
    return screen_rows


def get_response_label(response):
    """
    Получить текстовую метку респонса
    """
    if response == 'Total':
        return DAY_TOTAL
    elif ResponseCode.objects.filter(code=response).exists():
        return ResponseCode.objects.get(code=response).label
    else:
        return '{} - Неизвестный респонс'.format(str(response))


def get_table_headers(target_dates, selected_survey):
    """
    Получить шапки таблицы для экрана и экселя
    @return:
    """
    table_headers = [column if type(column) == str else dt.strftime(column, '%d.%m.%Y') for column in target_dates[-9:]]
    table_headers.insert(0, selected_survey)
    if SURVEY_TOTAL not in table_headers:
        table_headers.insert(1, SURVEY_TOTAL)
    return table_headers


def create_excel_stat(selected_survey, date_start, date_end, query_type):
    """Создать excel-файл статистики прозвона для скачивания пользователем"""
    sample_data, format_data = get_sample_data(selected_survey, query_type)
    target_dates = get_dates(date_start, date_end, sample_data)
    excel_rows = get_excel_rows(format_data, target_dates)
    wb = xl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1).value = selected_survey
    for col_num, col_label in enumerate(target_dates):
        if type(col_label) == str:
            ws.cell(row=1, column=col_num + 2).value = col_label
        else:
            ws.cell(row=1, column=col_num + 2).value = dt.strftime(col_label, '%d.%m.%Y')

    for row_num, resp_counts in enumerate(excel_rows):
        for label, counts in resp_counts.items():
            ws.cell(row=row_num + 2, column=1).value = label
            for col_num, count in enumerate(counts):
                ws.cell(row=row_num + 2, column=col_num + 2).value = count
    auto_fit_excel_columns(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    if query_type == 1:
        filename = 'Sample_count_{survey}.xlsx'.format(survey=translify(selected_survey).replace('.', '_'))
    elif query_type == 2:
        filename = 'Dialing_count_{survey}.xlsx'.format(survey=translify(selected_survey).replace('.', '_'))
    response['Content-Disposition'] = 'attachment; filename={}'.format(filename)
    wb.save(response)
    return response


def get_content(date_start, date_end, selected_survey, form, page_name, query_type):
    """
    Получить данные для страницы ЛК
    @param query_type: тип запрошенной инфы, 1 - из сэмпла, 2 - из ContactLog
    @type query_type: int
    """
    sample_data, format_data = get_sample_data(selected_survey, query_type)
    target_dates = get_dates(date_start, date_end, sample_data)
    screen_rows = get_screen_rows(format_data, target_dates)
    table_headers = get_table_headers(target_dates, selected_survey)
    dates_count = False
    if len(target_dates) > 11:
        dates_count = True
    render_content = {
        'form': form, 'headers': table_headers, 'rows': screen_rows, 'date_start': dt.strftime(date_start, '%Y-%m-%d'),
        'date_end': dt.strftime(date_end, '%Y-%m-%d'), 'page_name': page_name, 'dates_count': dates_count
    }
    return render_content


def get_view_content(request, page_name, query_type):
    survey_name, date_start, date_end, form, selected_survey = get_request_data(request)
    if selected_survey == 'Выберите проект':
        return cabinet_empty_render(request, form, page_name, 'no_survey', 'cabinet/dialing.html', date_start, date_end)
    else:
        render_content = get_content(date_start, date_end, selected_survey, form, page_name, query_type)
        if len(render_content['headers']) == 2:
            return cabinet_empty_render(
                request, form, page_name, 'no_work', 'cabinet/dialing.html', date_start, date_end
            )
        if request.POST.get("download_excel"):
            return create_excel_stat(selected_survey, date_start, date_end, query_type)
        return render(request, 'cabinet/dialing.html', render_content)


def get_stratification_content(request, page_name):
    from cabinet.forms import SurveyStatForm
    survey_name = request.POST.get("SurveyName")
    form = SurveyStatForm(initial={'SurveyName': survey_name})
    selected_survey = get_surveys(quota=True)[int(survey_name)][1]
    if selected_survey == 'Выберите проект':
        return cabinet_empty_render(request, form, page_name, 'no_survey', 'cabinet/stratification.html')
    else:
        strat_rows = get_startification(get_internal_survey(selected_survey))
        if request.POST.get("download_excel"):
            return create_excel_stratification(strat_rows, selected_survey)
        render_content = {'form': form, 'page_name': page_name, 'data': strat_rows}
        return render(request, 'cabinet/stratification.html', render_content)


def get_internal_survey(survey_name):
    """
    Получить объект внутреннего проекта
    """
    if Internal_Survey.objects.filter(cati_programm_name=survey_name).exists():
        return Internal_Survey.objects.get(cati_programm_name=survey_name)
    else:
        return Internal_Survey.objects.get(name=survey_name)


def get_startification(survey):
    strat_rows, quota_fields = get_stratification_map(survey)
    sample = get_stratification_data(survey.cati_programm_name, quota_fields)
    for enum, row in enumerate(strat_rows):
        quota_code = row[1].lower()
        if sample[row[0].lower()].dtype == 'int64':
            quota_code = int(quota_code)
        elif sample[row[0].lower()].dtype == 'float64':
            quota_code = float(quota_code)
        count = len(sample.loc[sample[row[0].lower()] == quota_code].index)
        if count > 0 and int(row[2]) > 0:
            strat_rows[enum].extend([count, int(round((count / int(row[2]) * 100), 0))])
        elif count > 0:
            strat_rows[enum].extend([count, 0])
        else:
            strat_rows[enum].extend([0, 0])
    return strat_rows


def get_stratification_map(survey):
    strat_rows = []
    quota_fields = set()
    for row in open(survey.stratification.path).readlines():
        strat_requisite = re.split(r'[:()]', row.replace('\n', ''))
        if strat_requisite[0]:
            strat_rows.append([
                strat_requisite[0].lower().replace('ttt', 't'),
                strat_requisite[1].lower().split('p')[0].replace('"', '').strip(),
                int(strat_requisite[2].lower().split('p')[0]), ' '.join(strat_requisite[3:]).strip()
            ])
            quota_fields.add(strat_requisite[0].lower().replace('ttt', 't'))
    return strat_rows, quota_fields


def get_stratification_data(survey_name, fields):
    fields = ','.join(fields)
    # Для мерсов нужно подсчитывать 18+59, для остальных - только 18
    if survey_name.lower() in ('mta2020', 'mts2020', '5star', '5stasm'):
        return pd.read_sql(
            'select {fields} from Sample{survey} where responsecode in (18, 59)'.format(
                fields=fields, survey=survey_name), nipo_connect()
        )
    else:
        return pd.read_sql(
            'select {fields} from Sample{survey} where responsecode = 18'.format(
                fields=fields, survey=survey_name), nipo_connect()
        )


def create_excel_stratification(strat_rows, selected_survey):
    wb = xl.Workbook()
    ws = wb.active
    for enum, cell in enumerate(('Описание квоты', 'Всего успешных', 'Размер квоты', 'Процент выполнения')):
        ws.cell(row=1, column=(enum + 1)).value = cell
    for enum, row in enumerate(strat_rows):
        ws.cell(row=enum + 2, column=1).value = row[3]
        ws.cell(row=enum + 2, column=2).value = row[4]
        ws.cell(row=enum + 2, column=3).value = int(row[2])
        ws.cell(row=enum + 2, column=4).value = '{val}%'.format(val=str(row[5]))
    auto_fit_excel_columns(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Stratification_{survey}.xlsx'.format(
        survey=translify(selected_survey).replace('.', '_')
    )
    wb.save(response)
    return response


def get_fw_monirot_content(request, page_name):
    survey_name, date_start, date_end, form, selected_survey = get_request_data(request)
    if selected_survey == 'Выберите проект':
        return cabinet_empty_render(request, form, page_name, 'no_survey', 'cabinet/dialing.html', date_start, date_end)
    else:
        data = get_fw_monitor_data(selected_survey, date_start, date_end)
        if len(data.index) == 0:
            return cabinet_empty_render(
                 request, form, page_name, 'no_inter_work', 'cabinet/interviewers_performance.html',
                 date_start, date_end=dt.now()
            )
        if request.POST.get("download_excel"):
            return create_fw_excel_stat(selected_survey, data)
        data = calculate_fw_monitor(data).values
        headers = [selected_survey, *FW_MONITOR_HEADERS]
        render_content = {
            'form': form, 'headers': headers, 'rows': data, 'date_start': dt.strftime(date_start, '%Y-%m-%d'),
            'date_end': dt.strftime(date_end, '%Y-%m-%d'), 'page_name': page_name
        }
        return render(request, 'cabinet/interviewers_performance.html', render_content)


def get_fw_monitor_data(survey_name, date_start, date_end):
    fields = ['contactdatetime', 'responsestatus', 'interviewernumber']
    fields.extend(CONTACT_LOG_TIME_COLUMNS)
    data = pd.read_sql(
        "select {fields} from ContactLog where surveyname = '{survey}' and interviewernumber > 0"
        "and ContactDateTime between '{date_start}' and '{date_end}'".format(
            fields=','.join(fields), survey=get_sample_name(survey_name).replace('Sample', ''), date_start=date_start,
            date_end=date_end), nipo_connect()
    )
    return data


def calculate_fw_monitor(data):
    data['date'] = data['contactdatetime'].dt.date
    data['spent_time'] = data[CONTACT_LOG_TIME_COLUMNS].sum(axis=1)
    data.drop([*CONTACT_LOG_TIME_COLUMNS, 'contactdatetime'], axis='columns', inplace=True)
    agregate_date = pd.DataFrame(
        {'date': sorted(data['date'].unique()),
         'numb_succesful': data['date'].loc[data['responsestatus'] == 18].value_counts(),
         'interviewer_numbers': data.groupby('date')['interviewernumber'].nunique(),
         'spent_time': data.groupby('date')['spent_time'].sum() / 3600,
         }
    ).fillna(0)
    agregate_date['perfomance'] = round(agregate_date['numb_succesful'] / agregate_date['spent_time'], 2)
    agregate_date['spent_time'] = round(agregate_date['spent_time'], 2)
    agregate_date['numb_succesful'] = agregate_date['numb_succesful'].astype('int')
    agregate_date.fillna('0', inplace=True)
    # Зафиксировать порядок колонок
    agregate_date = agregate_date[['date', 'numb_succesful', 'interviewer_numbers', 'spent_time', 'perfomance']]
    return agregate_date


def create_fw_excel_stat(selected_survey, agregate_date):
    wb = xl.Workbook()
    ws = wb.active
    agregate_date = calculate_fw_monitor(agregate_date)
    headers = [selected_survey, *FW_MONITOR_HEADERS]
    for col_num, col_label in enumerate(headers):
        ws.cell(row=1, column=(col_num + 1)).value = col_label
    for enum, row in enumerate(agregate_date.values):
        ws.cell(row=enum + 2, column=1).value = dt.strftime(row[0], '%d.%m.%Y')
        ws.cell(row=enum + 2, column=2).value = row[1]
        ws.cell(row=enum + 2, column=3).value = row[2]
        ws.cell(row=enum + 2, column=4).value = row[3]
        ws.cell(row=enum + 2, column=5).value = row[4]
    auto_fit_excel_columns(ws)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Interviewers_productive_{survey}.xlsx'.format(
        survey=translify(selected_survey).replace('.', '_')
    )
    wb.save(response)
    return response


def get_request_data(request):
    from cabinet.forms import SurveyStatForm
    survey_name = request.POST.get("SurveyName")
    date_start = dt.strptime(request.POST.get("date_start"), '%Y-%m-%d')
    date_end = dt.strptime(request.POST.get("date_end"), '%Y-%m-%d').replace(hour=23)
    form = SurveyStatForm(initial={'SurveyName': survey_name})
    selected_survey = get_surveys()[int(survey_name)][1]
    return survey_name, date_start, date_end, form, selected_survey
