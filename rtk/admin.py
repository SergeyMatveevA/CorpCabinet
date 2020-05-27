import openpyxl as xl

from django.contrib import admin
from django.http import HttpResponse

from rtk.consts import TOUCH_POINTS, RTK_REGIONS, COMPL_PROCESSING, COMPL_RESULT, STATE
from rtk.models import Complaint


class ComplaintAdmin(admin.ModelAdmin):
    fields = [
        'id', 'telephone', 'abonent_id', 'region', 'touch_point', 'translate_reason', 'translate_result',
        'translate_result_other', 'appointment_time', 'appointment_comment', 'processed_comment', 'ready_recommend',
        'negative_reason_rtk', 'cc_satisfaction', 'negative_reason_cc', 'state_compl',  'creation_date', 'change_date',
        'audio'
    ]

    list_display = [
        'id', 'telephone', 'translate_reason', 'state_compl', 'creation_date', 'change_date', 'region', 'touch_point'
    ]
    search_fields = ['id', 'telephone', 'translate_reason', 'state_compl', 'creation_date', 'change_date']
    list_filter = ['id', 'translate_reason', 'state_compl', 'creation_date', 'change_date', 'region', 'touch_point']

    readonly_fields = (
        'id', 'abonent_id', 'creation_date', 'ready_recommend', 'cc_satisfaction', 'region', 'touch_point', 'audio',
    )

    def export_tickets(modeladmin, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={filename}.xlsx'.format(
            filename=str(modeladmin).replace('.', '_').replace('Admin', '')
        )
        wb = xl.Workbook()
        ws = wb.active

        columns = (
            'id', 'telephone', 'abonent_id', 'region', 'touch_point', 'translate_reason', 'translate_result',
            'translate_result_other', 'appointment_time', 'appointment_comment', 'processed_comment', 'ready_recommend',
            'negative_reason_rtk', 'cc_satisfaction', 'negative_reason_cc', 'state_compl',  'creation_date',
            'change_date'
        )

        headers = (
            'Номер ТТ', 'Номер телефона', 'ID абонента', 'МРФ', 'Точка контакта', 'Причина перевода',
            'Причина перевода', 'Результат перевода', 'Желаемое дата-время перезвона', 'Комментарий к перезвону',
            'Комментарий к обработке', 'Готовность рекомендовать Ростелеком', 'Причина низкой оценки Ростелеком',
            'Удовлетворенность работой ', 'Причина низкой оценки работы ', 'Статус', 'Дата создания', 'Дата изменения'
        )
        for col_num, val in enumerate(headers):
            ws.cell(row=1, column=col_num + 1).value = val

        regions = dict(RTK_REGIONS)
        touch_points = dict(TOUCH_POINTS)
        compl_proc = dict(COMPL_PROCESSING)
        compl_result = dict(COMPL_RESULT)
        state = dict(STATE)
        for row_num, obj in enumerate(queryset.values_list(*columns)):
            for cell, val in zip(ws.iter_cols(max_col=len(obj), max_row=row_num + 2, min_row=row_num + 2), obj):
                if val:
                    if cell[0].column == 4:
                        val = regions[val]
                    elif cell[0].column == 5:
                        val = touch_points[val]
                    elif cell[0].column == 6:
                        val = compl_proc[val]
                    elif cell[0].column == 7:
                        val = compl_result[val]
                    elif cell[0].column == 16:
                        val = state[val]

                cell[0].value = val

        wb.save(response)
        return response

    export_tickets.short_description = 'Выгрузить трабл-тикеты'
    actions = [export_tickets]


admin.site.register(Complaint, ComplaintAdmin)