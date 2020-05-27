import random

from datetime import datetime as dt
from re import fullmatch

import openpyxl as xl

from openpyxl.utils.cell import coordinate_from_string, column_index_from_string

from arrangement.consts import ST_ERR, ST_FIN
from yourls.models import YourlsUrl, Yourls2Url
from yourls_interface.consts import ID_MAP, ANONYMOUS_ENGINE, ID_ALREADY_USED, YOURLS_DB_ERROR, ID_NOT_FIT_YOURLS
from yourls_interface.models import YourlsID


def generate_short_links(generation):
    """Сгенерировать короткие ссылки на движке yourls
    :param generation: Одна генерация
    :type generation: LinkGeneration
    """
    wb = xl.load_workbook(generation.links_file)
    sheet_name, links_column = wb.defined_names['link'].attr_text.split('!')
    ws = wb[sheet_name]
    links_column_number = column_index_from_string(coordinate_from_string(links_column)[0])
    if generation.self_id:
        ids_column = column_index_from_string(coordinate_from_string(wb.defined_names['id'].attr_text.split('!')[1])[0])
    row_number = 2
    links = {}

    while ws.cell(row=row_number, column=links_column_number).value:
        links[row_number] = {}
        links[row_number]['link'] = ws.cell(row=row_number, column=links_column_number).value
        if generation.self_id:
            links[row_number]['id'] = ws.cell(row=row_number, column=ids_column).value
        row_number += 1
    if generation.self_id:
        if check_id_validity(links):
            if check_id_existence(links, generation.engine):
                generation.state = ST_ERR
                generation.processed_comments = ID_ALREADY_USED
                generation.save()
                return
        else:
            generation.state = ST_ERR
            generation.processed_comments = ID_NOT_FIT_YOURLS
            generation.save()
            return

    elif not generation.self_id:
        assign_id_to_links(links, generation.engine)

    ws.insert_cols(links_column_number + 1)
    ws.cell(row=1, column=links_column_number + 1).value = 'Короткая ссылка'
    for link in links:
        ws.cell(row=link, column=(links_column_number + 1)).value = \
            r'{eng_name}/{id}'.format(eng_name=generation.engine, id=links[link]['id'])
    wb.save(generation.links_file.path)
    if push_links_to_yourls(links, generation.engine, generation.title):
        generation.state = ST_FIN
    else:
        generation.state = ST_ERR
        generation.processed_comments = YOURLS_DB_ERROR
    generation.save()


def ids_gen():
    """Сгенерировать 100.000 случайных комбинаций букв и цифр длинной 4 символа"""
    new_ids = set()
    while len(new_ids) < 100000:
        new_ids.add(''.join([ID_MAP[random.randint(1, 32)] for x in range(0, 4)]))
    YourlsID.objects.bulk_create([YourlsID(yourls_id=id) for id in new_ids], ignore_conflicts=True)


def assign_id_to_links(links, engine):
    """Присвоить ссылкам сгенерированные id
    :type links: dict {row: {'link'; link, 'id': id}}
    :param engine: Движок yourls на котором будет короткая ссылка
    :type engine: LinkGeneration.engine
    """
    if engine == ANONYMOUS_ENGINE:
        filter_cond = {'use_in_mbr_engine': False, 'id_type': 1}
    else:
        filter_cond = {'use_in_general_engine': False, 'id_type': 1}
    while YourlsID.objects.filter(**filter_cond).count() < len(links):
        ids_gen()
    system_id = YourlsID.objects.filter(**filter_cond)[:len(links)]
    for number, id in enumerate(system_id):
        links[number + 2]['id'] = id.yourls_id
    if engine == ANONYMOUS_ENGINE:
        YourlsID.objects.filter(yourls_id__in=system_id).update(use_in_mbr_engine=True)
    else:
        YourlsID.objects.filter(yourls_id__in=system_id).update(use_in_general_engine=True)


def check_id_existence(links, engine):
    """
    Проверить id из файла на повторное использование и записать их в базу если id оригинальные
    :type links: dict {row: {'link'; link, 'id': id}}
    """
    ids = [links[row]['id'] for row in links]
    if YourlsID.objects.filter(yourls_id__in=ids).exists():
        return True

    mbr_eng = False
    gen_eng = False
    if engine == ANONYMOUS_ENGINE:
        mbr_eng = True
    else:
        gen_eng = True
    new_ids = [
        YourlsID(yourls_id=id.lower(), id_type=2, use_in_general_engine=gen_eng, use_in_mbr_engine=mbr_eng)
        for id in ids
    ]
    YourlsID.objects.bulk_create(new_ids)
    return False


def push_links_to_yourls(links, engine, title):
    """Записать в БД Yourls новые шорт-линки
    :type links: dict {row: {'link'; link, 'id': id}}
    :param engine: Движок yourls на котором будет короткая ссылка
    :type engine: LinkGeneration.engine
    :param title: Подпись страницы для предзагрузки в мессенджерах
    :type title: str
    """
    if engine == ANONYMOUS_ENGINE:
        model = YourlsUrl
    else:
        model = Yourls2Url
    new_shortlinks = []
    for link in links:
        new_shortlinks.append(
            model(
                keyword=links[link]['id'], url=links[link]['link'], title=title, timestamp=dt.now(),
                ip='10.165.99.99', clicks=0
            )
        )
    model.objects.bulk_create(new_shortlinks)
    return True


def check_id_validity(links):
    """
    Проверить валидность кастомных id к правилам Yourls
    :type links: dict {row: {'link'; link, 'id': id}}
    """
    for link in links:
        if not fullmatch(r'^[a-zA-Z0-9]+$', links[link]['id']):
            return False
    return True
