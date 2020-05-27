import openpyxl as xl

from datetime import datetime as dt

from arrangement.consts import ST_FIN


def calculate_interviews(calculate):
    """
    Рассчитать количество интервью на интервьюера в разрезе по выборкам
    """
    wb = xl.load_workbook(calculate.quotas_file.path, data_only=True)
    interviewers = {}
    samples = {'main': [], 'additional': []}
    for sheet in wb.get_sheet_names():
        if 'ячейка' in sheet.lower():
            samples['main'].append(sheet)
            row_number = 9
            quota_size = 15
        elif 'доп' in sheet.lower():
            samples['additional'].append(sheet)
            row_number = 12
            quota_size = 4
        else:
            continue
        calculate_sample(wb[sheet], interviewers, sheet, row_number, quota_size)
    count_sheet_name = 'Сводная {date}.{month}.{year}'.format(
        date=str(dt.now().day), month=str(dt.now().month), year=str(dt.now().year)
    )
    if count_sheet_name in wb.get_sheet_names():
        wb.remove_sheet(wb[count_sheet_name])
    wb.create_sheet(count_sheet_name, 0)
    ws = wb[count_sheet_name]
    col_number = 2
    sample_map = {}
    for s_type in (sorted(samples['main']), sorted(samples['additional'])):
        for sample in s_type:
            ws.cell(row=1, column=col_number).value = sample
            sample_map[sample] = col_number
            col_number += 1
    row_number = 2
    for inter in interviewers:
        ws.cell(row=row_number, column=1).value = inter
        for sample in interviewers[inter]:
            ws.cell(row=row_number, column=sample_map[sample]).value = interviewers[inter][sample]
        row_number += 1
    wb.save(calculate.quotas_file.path)
    calculate.state = ST_FIN
    return calculate


def calculate_sample(ws, interviewers, sample_name, row_number, quota_column):
    """Посчитать количество интервью в одной выборке"""
    while ws.cell(row=row_number, column=quota_column).value:
        inter = ws.cell(row=row_number, column=2).value
        if not inter:
            row_number += 1
            continue
        if inter not in interviewers:
            interviewers[inter] = {}
        if sample_name not in interviewers[inter]:
            interviewers[inter][sample_name] = 0
        interviewers[inter][sample_name] += int(ws.cell(row=row_number, column=quota_column).value)
        row_number += 1
    return interviewers
