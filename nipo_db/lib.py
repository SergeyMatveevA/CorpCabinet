import openpyxl as xl

from django.http import HttpResponse


def export_sample(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={filename}.xlsx'.format(
        filename=str(modeladmin).replace('.', '_').replace('Admin', '')
    )
    wb = xl.Workbook()
    ws = wb.active

    columns = [field.name for field in queryset[0]._meta.get_fields()]
    for col_num in range(len(columns)):
        ws.cell(row=1, column=col_num + 1).value = columns[col_num]

    for row_num, obj in enumerate(queryset.values_list()):
        for cell, val in zip(ws.iter_cols(max_col=len(obj), max_row=row_num + 2, min_row=row_num + 2), obj):
            cell[0].value = val

    wb.save(response)
    return response


export_sample.short_description = 'Выгрузить строки сэмпла'
