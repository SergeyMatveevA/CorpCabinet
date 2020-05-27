from django.contrib import admin
from django.http import HttpResponse

import openpyxl as xl

from nipo_db.lib import export_sample
from nipo_db.models import Sampledkcsias, Sampledkcsisa, Samplemta2020, Sample5Star, Sample5Stasm

STANDARD_FIELDS = [
    'interviewnumber', 'responsecode', 'status', 'numberofcontacts', 'id', 'email', 'contacttime', 'channel',
    'initialchannel', 'systemdata', 'appointmenttime', 'appointmentname', 'telephonenumber', 'interviewernumber',
    'targetinterviewer', 'showdisplayfields', 'displayfield1', 'displayfield2', 'displayfield3', 'displayfield4',
    'secondphonenumber', 'timedifference', 'callintervalbegin', 'callintervalend', 'language', 'extradata'
]

STANDARD_SEARCH = ['interviewnumber', 'telephonenumber', 'secondphonenumber']
STANDARD_FILTER = ['responsecode', 'status', 'numberofcontacts']
STANDARD_LIST_DISPLAY = [
    'interviewnumber', 'responsecode', 'status', 'numberofcontacts', 'email', 'contacttime', 'systemdata',
    'appointmenttime', 'appointmentname', 'telephonenumber', 'interviewernumber', 'secondphonenumber', 'timedifference',
    'callintervalbegin', 'callintervalend'
]


class DkcsisaAdmin(admin.ModelAdmin):
    survey_fields = [
        'dealer_code', 'dealer_name', 'dateobs', 'remontzakaz', 'vin', 'vin_am', 'vehicle_year', 'client_type',
        'region', 'name', 'screenoutquestion', 'dealercodewave', 'wave'
    ]
    fields = STANDARD_FIELDS.extend(survey_fields)
    STANDARD_LIST_DISPLAY.extend(survey_fields)
    list_display = STANDARD_LIST_DISPLAY
    STANDARD_SEARCH.extend(survey_fields)
    search_fields = STANDARD_SEARCH
    list_filter = ['dealer_code', 'wave', 'responsecode', 'status', 'numberofcontacts']
    actions = [export_sample]


class DkcsiasAdmin(admin.ModelAdmin):
    survey_fields = [
        'dealer_code', 'dealer_name', 'dateobs', 'remontzakaz', 'vin', 'vin_am', 'vehicle_year', 'client_type',
        'region', 'name', 'screenoutquestion', 'dealercodewave', 'wave'
    ]
    fields = STANDARD_FIELDS.extend(survey_fields)
    STANDARD_LIST_DISPLAY.extend(survey_fields)
    list_display = STANDARD_LIST_DISPLAY
    STANDARD_SEARCH.extend(survey_fields)
    search_fields = STANDARD_SEARCH
    list_filter = ['dealer_code', 'wave', 'responsecode', 'status', 'numberofcontacts']
    actions = [export_sample]


class Mta2020Admin(admin.ModelAdmin):
    fields_names = [field.name for field in Samplemta2020._meta.get_fields() if field.name != 'suspendimage']
    fields = fields_names
    list_display = fields_names
    search_fields = ['name', 'remont_zakaz', 'vin', 'master_name', 'dealername', 'dealercodewave', 'telephonenumber']
    list_filter = ['wave', 'dealer_code', 'responsecode']

    def export_sample(modeladmin, request, queryset):
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename={filename}.xlsx'.format(filename='response_report')

        wb = xl.Workbook()
        ws = wb.active
        response_labels = {
            0: 'Фреш', 1: 'Нет ответа', 3: 'Занято', 4: 'Недоступен', 5: 'Неправильный номер', 6: 'Договоренность',
            7: 'Договоренность', 8: 'Отказ', 17: 'Уже опрашивали', 18: 'Успешное', 20: 'Недоступен', 21: 'Фреш',
            28: 'Прерванное', 29: 'Прерванное', 41: 'Недоступен', 42: 'Недоступен', 44: 'Нет ответа',
            45: 'Автоответчик', 51: 'Прерванное', 52: 'Отказ', 19: 'скринаут', 43: 'Автоответчик', 59: 'Успешное'
        }
        columns = [field.name for field in Samplemta2020._meta.get_fields() if field.name != 'suspendimage']
        wave_column = 0
        for col_num in range(len(columns)):
            if columns[col_num] == 'wave' and wave_column == 0:
                wave_column = col_num + 1
            ws.cell(row=1, column=col_num + 1).value = columns[col_num]

        for row_num, obj in enumerate(queryset.values_list(*tuple(columns))):
            for cell, val in zip(ws.iter_cols(max_col=len(obj), max_row=row_num + 2, min_row=row_num + 2), obj):
                if cell[0].column == 2:
                    if val in response_labels:
                        val = response_labels[val]
                    else:
                        val = 'Неизвестный респонс ({number})'.format(number=str(val))
                cell[0].value = val
        # Добавим отдельной процедурой изменение номера интервью по Лёшиному алгоритму
        row_num = 2
        while ws.cell(row=row_num, column=1).value:
            ws.cell(row=row_num, column=1).value = \
                int(ws.cell(row=row_num, column=wave_column).value[:2]) * 100000 + \
                ws.cell(row=row_num, column=1).value + 10000
            row_num += 1

        wb.save(response)
        return response
    export_sample.short_description = 'Выгрузить отчёт по респонс-кодам'
    actions = [export_sample]


class Star5Admin(admin.ModelAdmin):
    fields_names = [field.name for field in Sample5Star._meta.get_fields() if field.name != 'suspendimage']
    fields = fields_names
    list_display = fields_names
    search_fields = ['name', 'remont_zakaz', 'vin', 'dealername', 'telephonenumber']
    list_filter = ['wave', 'dealercode', 'dealercodewave', 'responsecode']
    actions = [export_sample]


class Stasm5Admin(admin.ModelAdmin):
    fields_names = sorted([field.name for field in Sample5Stasm._meta.get_fields() if field.name != 'suspendimage'])
    fields = fields_names
    list_display = fields_names
    search_fields = ['name', 'remont_zakaz', 'vin', 'dealername', 'telephonenumber']
    list_filter = ['wave', 'origin_dealer_code', 'dealercodewave', 'responsecode']
    actions = [export_sample]


admin.site.register(Sampledkcsias, DkcsiasAdmin)
admin.site.register(Sampledkcsisa, DkcsisaAdmin)
admin.site.register(Samplemta2020, Mta2020Admin)
admin.site.register(Sample5Star, Star5Admin)
admin.site.register(Sample5Stasm, Stasm5Admin)
