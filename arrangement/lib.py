import datetime as dt
import openpyxl as xl
import os
import re
import smtplib
import shutil

import numpy as np
import pandas as pd

from datetime import timedelta, datetime
from email.mime.text import MIMEText

from django.apps import apps
from django.db.models import Q
from django.http import HttpResponse
from sqlalchemy import create_engine

from arrangement.consts import STRATS_PATH, MM_QUOTA, MS_QUOTA, SL_QUOTA, CN_QUOTA, base_plug
from nipo_db.models import Contactlog, Survey as nipo_survey

def nipo_connect():
    """Создать подключене к nipo"""
    return create_engine('mssql+pymssql://anonymous:anonymous@192.168.50.99/anonymous', echo=False)


def self_connect():
    """Создать подключение к бд проекта"""
    return create_engine('postgresql+psycopg2://sv:SV@127.0.0.1/sv')


def check_file_headers(template, exemplar):
    """
    Проверяет соответствие и порядок колонок, возвращает true\false
    :type template: str
    :param exemplar: str
    :rtype: bool
    """
    headers = []
    for book in (template, exemplar):
        ws = xl.load_workbook(book).active
        book_headers = []
        for column in range(1, ws.max_column + 1):
            book_headers.append(ws.cell(row=1, column=column).value)
        headers.append(book_headers)
    if headers[0] != headers[1]:
        return False
    return True


def export_xlsx(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={filename}.xlsx'.format(
        filename=str(modeladmin).replace('.', '_').replace('Admin', '')
    )
    wb = xl.Workbook()
    ws = wb.active

    columns = [field.name for field in queryset[0]._meta.get_fields()]
    for col_num in range(len(columns)):
        ws.cell(row=1, column=col_num + 1).value = columns[col_num]

    for row_num, obj in enumerate(queryset.values_list(columns)):
        for cell, val in zip(ws.iter_cols(max_col=len(obj), max_row=row_num + 2, min_row=row_num + 2), obj):
            cell[0].value = val

    wb.save(response)
    return response

export_xlsx.short_description = 'Выгрузить xlsx'


def export_stratification(modeladmin, request, queryset):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={filename}.xlsx'.format(
        filename=str(modeladmin).replace('.', '_').replace('Admin', '')
    )
    wb = xl.Workbook()
    ws = wb.active

    columns = (
        'numbers_of_complited', 'percentage_of_target', 'target', 'numbers_of_eliminated', 'numbers_of_fresh',
        'numbers_of_no_answered', 'numbers_of_appointments', 'description'
    )
    for col_num, val in enumerate(columns):
        ws.cell(row=1, column=col_num + 1).value = val

    for row_num, obj in enumerate(queryset.values_list(*columns)):
        for cell, val in zip(ws.iter_cols(max_col=len(obj), max_row=row_num + 2, min_row=row_num + 2), obj):
            cell[0].value = val

    wb.save(response)
    return response


export_stratification.short_description = 'Выгрузить стратификацию'


def stratification_reload(survey, strat_address):
    """
    Перезагрузить стратифкацию проекта.
    :param survey: Кати-проект
    :type survey: arrangement.models.Survey
    :param strat_address: адрес файла стратификации на сервере
    :type strat_address: str
    """
    if survey.stratification:
        os.remove(survey.stratification.path)
    strat_path = os.path.join(STRATS_PATH, '{survey}s'.format(survey=survey.cati_programm_name.title()))
    shutil.copy2(strat_address, strat_path)
    survey.stratification = strat_path
    survey.stratification_update_date = dt.datetime.fromtimestamp(os.stat(strat_path).st_mtime)
    survey.save()


def read_quota_fields(survey):
    """
    Считать из файла стратифкации квотные поля
    :param survey: Кати-проект
    :type survey: arrangement.models.Survey
    :return quotas: словарь вида {ПолеВБД, КодВПолеБД, РазмерКвоты, ОписаниеКвоты}
    """

    quotas = []
    with open(survey.stratification.path, 'r') as strat_file:
        for row_number, line in enumerate(strat_file.readlines()):
            target = 0
            if re.search(r'\((  )?\d+\)', line):
                target = int(re.search(r'\((  )?\d+\)', line).group()[1:-1])
            quotas.append({
                'row_number': row_number,
                'field': re.search(r'\w*\b', line).group(),
                'field_val': re.search(r':\w*\b', line).group()[1:],
                'target': target,
                'description': line.split(') ', maxsplit=1)[1:][0].replace('\n', '')
            })
    return quotas


def stratification_fill(survey):
    """
    Заполнить стратифкацию проекта
    :param survey: Кати-проект
    :type survey: arrangement.models.Survey
    """
    quota_req = read_quota_fields(survey)
    for quota in quota_req:
        model = apps.get_model(survey.cati_programm_name.lower(), 'Stratification')
        model.objects.create(**quota)


def stratification_calc(survey):
    """
    Посчитать показатели стратифации
    :param survey: Кати-проект
    :type survey: arrangement.models.Survey
    """
    stats = apps.get_model(survey.cati_programm_name.lower(), 'Stratification').objects.all()
    sample = apps.get_model(survey.cati_programm_name.lower(), 'Sample').objects.all()
    for row in stats:
        field_and_val = {row.field.lower(): row.field_val}
        row.numbers_of_complited = sample.filter(responsecode=18, **field_and_val).count()
        row.numbers_of_fresh = sample.filter(responsecode=0, **field_and_val).count()
        row.numbers_of_appointments = sample.filter(responsecode__in=[6,7], **field_and_val).count()
        row.numbers_of_no_answered = sample.filter(
            responsecode__in=[1, 2, 3, 20, 28, 29, 41, 42, 43, 44, 45], **field_and_val
        ).count()
        row.numbers_of_eliminated = sample.filter(
            responsecode__in=[4, 5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 19, 21, 51, 52], **field_and_val
        ).count()
        if row.target == 0:
            row.percentage_of_target = '0%'
        else:
            row.percentage_of_target = '{perc}%'.format(perc=int(row.numbers_of_complited / row.target * 100))
        row.save()


def get_quota_map(survey):
    """
    :param survey: Кати-проект
    :type survey: arrangement.models.Survey
    """
    # Получить не использованные записи контактов длля получения значений заливаемых в сэмпл
    fresh_rows = apps.get_model(survey.cati_programm_name.lower(), 'Sample').objects.filter(numberofcontacts=0)

    # Из стратификации получим набор квотных полей и их значения
    quota_rows = apps.get_model(survey.cati_programm_name.lower(), 'Stratification').objects.exclude(target=0)\
        .exclude(percentage_of_target='100%')
    quota_fields = {}
    for row in quota_rows:
        if row.field.lower() not in quota_fields:
            quota_fields[row.field.lower()] = []
        quota_fields[row.field.lower()].append(row.field_val)
    # Отфильтруем только заливаемые в базу поля и значения среди незакрытых
    load_fields = {}
    for field, vals in quota_fields.items():
        for val in vals:
            predicate = {field: val}
            if fresh_rows.filter(**predicate).count() > 0:
                if field not in load_fields:
                    load_fields[field] = {'values': [], 'relations': {}}
                load_fields[field]['values'].append(val)

    load_fields = get_relations(load_fields, fresh_rows)
    # Найдём квоты верхнего уровня, чтобы раздлеить базу на сегменты
    senior_quotas = []
    slave_quotas = set()
    for field, data in load_fields.items():
        for rel_fields, relation in data['relations'].items():
            if relation == MS_QUOTA:
                slave_quotas.add(field)
        if field not in slave_quotas:
            senior_quotas.append(field)


def get_relations(quotas, base):
    """
    Определить тип связзей между полями квот по базе
    :param quotas: Открытые и заливаемые квоты
    :type quotas: dict
    :param base: Стандартная модель базы кати-опроса
    :type base: app.Sample
    :return: dict
    """
    for check_field in quotas:
        for field in quotas:
            if check_field != field and not quotas[check_field]['relations']:
                for check_val in quotas[check_field]['values']:
                    check_predicate = {check_field: check_val}
                    # Если одному значению проверяемого поля могут соответствовать больше одного значения второго поля,
                    # значит это либо связь вида хозяин => раб, либо многие ко многим
                    if base.filter(**check_predicate).distinct(field).count() > 1:
                        # Проверить есть ли для какого-либо значения поля с которым сверяем более одного значения поля
                        # которое проверяем (check_field)
                        for val in quotas[field]['values']:
                            predicate = {field: val}
                            if base.filter(**predicate).distinct(check_field).count() > 1:
                                quotas[check_field]['relations'][field] = MM_QUOTA
                                quotas[field]['relations'][check_field] = MM_QUOTA
                        # Если связь многие ко многим не была обнаружена - значит это связь хозяин-раб
                        if field not in quotas[check_field]['relations']:
                            quotas[check_field]['relations'][field] = SL_QUOTA
                            quotas[field]['relations'][check_field] = MS_QUOTA
                # Если после проверки свех значенй поля на связи мм/хр они не найдены, то это либо раб-хозяин,
                # либо связи конкурентная связь. Проверить на наличие раб-хозяин
                if field not in quotas[check_field]['relations']:
                    for check_val in quotas[check_field]['values']:
                        check_predicate = {check_field: check_val}
                        if base.filter(**check_predicate).distinct(field).count() == 1:
                            quotas[check_field]['relations'][field] = MS_QUOTA
                            quotas[field]['relations'][check_field] = SL_QUOTA
                    # Если не найдена связь раб-хозяин - связь конкурентная (= отсутствие связи)
                    if field not in quotas[check_field]['relations']:
                        quotas[check_field]['relations'][field] = CN_QUOTA
                        quotas[field]['relations'][check_field] = CN_QUOTA
    return quotas

export_xlsx.short_description = 'Выгрузить xlsx'


def check_duplicate(sample, checks):
    """
    Проверить дата-фрейм на налиие дублей по заданным типам и полям
    :param sample: данные для проверки
    :type sample: pd.DataFrame
    :param checks: проверяемые кортежи полей
    :type checks: list
    """
    if 'ResponseCode' not in sample:
        sample['ResponseCode'] = ''

    for check in checks:
        base_requisites = {}
        for row in sample.iterrows():
            if row[1]['ResponseCode'] == '':
                row_requisites = set()
                for field in check:
                    row_requisites.add(row[1][field])
                for requisite in row_requisites:
                    if requisite not in base_plug and not requisite is np.NaN:
                        if requisite not in base_requisites:
                            base_requisites[requisite] = [row[0]]
                        else:
                            base_requisites[requisite].append(row[0])
        check_type = ''
        if 'тел' in check[0].lower():
            check_type = ' по номеру телефона'
        elif 'mail' in check[0].lower():
            check_type = ' по электронной почте'
        elif 'vin' in check[0].lower():
            check_type = ' по vin-коду'
        rows_for_proc = set()
        for requisite in base_requisites:
            if len(base_requisites[requisite]) > 1:
                for row in base_requisites[requisite][1:]:
                    rows_for_proc.add(row)
        sample.loc[rows_for_proc, 'ResponseCode'] = 'Дубль внутри базы{check_type}'.format(check_type=check_type)


def check_new_surveys():
    """
    Проверить по контакт-логу наличие новых проектов
    """
    from arrangement.models import Survey
    start_date = dt.datetime.now() - timedelta(minutes=30)
    survey_names = [
        (s['surveyname'], s['server'])
        for s in Contactlog.objects.filter(
            ~Q(surveyname__in=('Test4', 'Ncbstst', '35109f', '55639f')), contactdatetime__gte=start_date
        ).values('surveyname', 'server').distinct()
    ]
    for survey in survey_names:
        if not Survey.objects.filter(cati_programm_name=survey[0]).exists():
            Survey.objects.create(
                cati_programm_name=survey[0], server_name=survey[1], last_sample_activity=dt.datetime.now()
            )


def fill_survey_names():
    """
    Заполнить имена проектов
    """
    from arrangement.models import Survey
    anonymous = Survey.objects.filter(
        name__isnull=True, last_sample_activity__gt=(dt.datetime.now() - timedelta(hours=24))
    )
    for survey in anonymous:
        if nipo_survey.objects.filter(surveyname=survey.cati_programm_name, excelsheet__isnull=False).exists():
            survey.name = nipo_survey.objects.get(surveyname=survey.cati_programm_name).excelsheet
            survey.save()


def auto_fit_excel_columns(worksheet):
    """Сделать автоширину колонок Excel-workbook"""
    for col in worksheet.iter_cols():
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[column].width = adjusted_width


def update_sample_activity():
    """
    Обновить даты активности сэмплов
    """
    # Чтобы избежать циклического импорта lib/models arrangement
    internal_survey = apps.get_model('arrangement', 'Survey')
    active_surveys = Contactlog.objects.filter(
        contactdatetime__gte=(datetime.now()-timedelta(minutes=15))
    ).values_list('surveyname', flat=True).distinct()
    for survey in active_surveys:
        if internal_survey.objects.filter(cati_programm_name=survey).exists():
            updated_survey = internal_survey.objects.get(cati_programm_name=survey)
            updated_survey.last_sample_activity = datetime.now()
            updated_survey.save()


def send_mail(recipients, message_text, subject):
    """
    :param recipient: почта нужного филиала
    """
    sender = 'anonymous@anonymous.ru'
    server = 'anonymous.ru:2525'
    msg = MIMEText(message_text.encode('utf-8'), _charset='utf-8', _subtype='html')
    msg['Subject'] = subject
    msg['From'] = sender
    msg['To'] = ','.join(recipients)

    s = smtplib.SMTP(server)
    s.login('anonymous@anonymous.ru', 'anonymous')
    s.sendmail(sender, recipients, msg.as_string())
    s.quit()
